from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import Select
import re
import openpyxl

# 반복문 이용하여 리스트에 있는 공백 제거
def del_Blink(List):
    deleted_List = []
    for i in List:
        if i != '':
            deleted_List.append(i)
    return deleted_List

# 정규표현식을 이용하여 문자데이터 정제
def clean_text(List):
    cleaned_List = []
    for i in List:
        i = re.sub('[\{\}\[\]\/?;:|*~`!^\-_+<>@\#\\n$%&\\\=\'\"]''', '', i)
        cleaned_List.append(i)
    return cleaned_List


path = r'C:\Users\Kim Chanwoo\Desktop\Programming\chromedriver.exe'
driver = webdriver.Chrome(path)
# 2020년 1학기 시간표 기준
driver.get('https://wis.hufs.ac.kr/src08/jsp/lecture/LECTURE2020L.jsp')


select_campus = input('캠퍼스 입력 :')


if '서울' in select_campus:
    campus = driver.find_element_by_css_selector('body > div > form > div.table.write.margin_top30 > table > tbody > tr:nth-child(3) > td > label:nth-child(2)')
    campus.click()

    search_major = input('검색할 학과 :')
    for num in range(1, 72):
        majors = driver.find_element_by_xpath('/html/body/div/form/div[1]/table/tbody/tr[4]/td/div/select/option[{}]'.format(num))
        if search_major in majors.text:
            majors.click()

elif '글로벌' in select_campus:
    campus = driver.find_element_by_css_selector('body > div > form > div.table.write.margin_top30 > table > tbody > tr:nth-child(3) > td > label:nth-child(4)')
    campus.click()

    search_major = input('검색할 학과 :')
    for num in range(1, 64):
        majors = driver.find_element_by_xpath('/html/body/div/form/div[1]/table/tbody/tr[4]/td/div/select/option[{}]'.format(num))
        if search_major in majors.text:
            majors.click()


source = driver.page_source
bs = BeautifulSoup(source, 'html.parser')
subjects = bs.findAll('font', {'class' : 'txt_navy'})
# 동명 td 태그가 많아 select 이용하여 12번째 index 크롤링
pros = bs.select('td:nth-of-type(12)')


# 태그 제거, 순수 텍스트 추출
sub_List = []
pros_List = []


for sub in subjects:
    sub_List.append(sub.text.strip())

for pro in pros:
    pro = pro.text
    pro = ' '.join(pro.split())
    pros_List.append(pro)

sub_List = clean_text(sub_List)
sub_List = del_Blink(sub_List)
pros_List = clean_text(pros_List)

sub_prof = list(zip(sub_List, pros_List))

# 엑셀 저장
Info_xlsx = openpyxl.Workbook()
sheet1 = Info_xlsx.active
sheet1.title = input('학과명 :')
sheet1['B1'] = '과목명'
sheet1['C1'] = '담당 교수'

# 행 : row / 열 : column
for index in range(len(sub_prof)):
    sheet1.cell(row=index+2, column=1).value = index+1
    sheet1.cell(row=index+2, column=2).value = sub_prof[index][0]
    sheet1.cell(row=index+2, column=3).value = sub_prof[index][1]

Info_xlsx.save('과목 정보' + '(' + input('파일명: ') + ').xlsx')

driver.quit()

